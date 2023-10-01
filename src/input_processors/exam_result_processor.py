import os
import openpyxl
import pandas as pd
import numpy as np



def load_exam(file_path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    for i in range(12,52):
        ws.cell(row=2,column=i).value = f"{ws.cell(row=2,column=i).value} [{ws.cell(row=1,column=(i-(i-2)%5)).value}]"
    data = np.array(list(ws.values))
    cols = data[1]
    data = data[data[:,2]=="Kutno"]

    return pd.DataFrame(data,columns=cols)


def process_files_in_directory(dirname: str) -> pd.DataFrame:
    reports = None
    for filename in os.listdir(dirname):
        if filename.startswith("Wyniki_"):
            with open("../Kutno_HackSQL/" + filename, "r") as f:
                result = load_exam("../Kutno_HackSQL/" + filename)
                result['Year'] = int(filename[-9:-5])
                if reports is None:
                    reports = result
                else:
                    reports = pd.concat([reports,result])

    return reports